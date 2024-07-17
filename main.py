import pandas as pd
import re
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Функция для исправления орфографических ошибок в отдельных словах
def correct_word(word, correct_list):
    word_lower = word.lower()
    best_match = process.extractOne(word_lower, correct_list, scorer=fuzz.ratio)
    if best_match and best_match[1] >= 80:  # Порог уверенности 80%
        return best_match[0]
    return word  # Возвращаем исходное слово, если совпадение недостаточно хорошее

# Функция для восстановления знаков препинания
def restore_punctuation(original, corrected):
    punctuations = re.findall(r'\W+', original)
    words = re.findall(r'\w+', corrected)
    result = []
    for i in range(max(len(words), len(punctuations))):
        if i < len(words):
            result.append(words[i])
        if i < len(punctuations):
            result.append(punctuations[i])
    return ''.join(result)

# Функция для исправления ошибок в строках
def correct_spelling(value, correct_list):
    if isinstance(value, str):  # Проверяем, что значение является строкой
        words = re.findall(r'\w+', value)
        corrected_words = []
        for word in words:
            corrected_word = correct_word(word, correct_list)
            if word.lower() != corrected_word.lower() and fuzz.ratio(word.lower(), corrected_word) <= 80:
                return value  # Возвращаем исходное значение, если слово не найдено в словаре с приемлемой точностью
            corrected_words.append(corrected_word.capitalize() if word.istitle() else corrected_word)
        return restore_punctuation(value, ' '.join(corrected_words))
    return value  # Возвращаем исходное значение, если оно не строка

# Функция для исправления ошибок в заданных столбцах
def correct_columns(df, columns, correct_list):
    for column in columns:
        df[column] = df[column].apply(lambda x: correct_spelling(x, correct_list))
    return df

# Загружаем эталонный список слов из файла "dataset.xlsx"
standard_file = 'dataset.xlsx'
standard_df = pd.read_excel(standard_file)

# Извлекаем уникальные слова из столбца "input" и "output"
correct_words = set()
for profession in standard_df['input']:
    if isinstance(profession, str):
        words = re.findall(r'\w+', profession)
        correct_words.update(words)
for profession in standard_df['output']:
    if isinstance(profession, str):
        words = re.findall(r'\w+', profession)
        correct_words.update(words)

# Добавляем слова из столбца "Наименование" в файлах "Должности_оригинал" и "Профессии_оригинал"
original_files = ['Должности_оригинал.xlsx', 'Профессии_оригинал.xlsx']
for file in original_files:
    original_df = pd.read_excel(file)
    for name in original_df['Наименование']:
        if isinstance(name, str):
            words = re.findall(r'\w+', name)
            correct_words.update(words)

# Исключаем числовые значения и последовательности только из заглавных букв, также удаляем дубликаты
correct_words = list({word.lower() for word in correct_words if not word.isdigit() and not word.isupper()})

# Сохранение словаря эталонных слов в файл
dictionary_file = 'Эталонный словарь.xlsx'
df_dictionary = pd.DataFrame(correct_words, columns=['Words'])
df_dictionary.to_excel(dictionary_file, index=False)
print(f"Словарь эталонных слов сохранен как {dictionary_file}")

# Обработка файла "Потребность персонала"
input_file1 = 'Потребность персонала.xlsx'
df1 = pd.read_excel(input_file1)
df1['Исходная Профессия'] = df1['Профессия']  # Добавляем столбец с исходными вариантами
df1_corrected = correct_columns(df1.copy(), ['Профессия'], correct_words)

# Сохранение результата в новый файл
output_file1 = 'Потребность персонала исправленный.xlsx'
df1_corrected.to_excel(output_file1, index=False)

# Обработка файла "ЦЗН"
input_file2 = 'ЦЗН.xlsx'
df2 = pd.read_excel(input_file2)
df2['Исходная Должность'] = df2['Должность']  # Добавляем столбец с исходными вариантами
df2['Исходная Специальность'] = df2['Специальность']  # Добавляем столбец с исходными вариантами
df2_corrected = correct_columns(df2.copy(), ['Должность', 'Специальность'], correct_words)

# Сохранение результата в новый файл
output_file2 = 'ЦЗН исправленный.xlsx'
df2_corrected.to_excel(output_file2, index=False)

# Функция для раскраски ячеек и восстановления исходных значений
def color_and_restore_cells(workbook, worksheet, df_original, columns):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    col_idx = {cell.value: cell.column for cell in worksheet[1]}

    for column in columns:
        original_col = 'Исходная ' + column
        for row in range(2, len(df_original) + 2):
            original_value = str(df_original.loc[row - 2, original_col])
            corrected_value = str(df_original.loc[row - 2, column])

            original_clean = re.sub(r'\W+', '', original_value.lower())
            corrected_clean = re.sub(r'\W+', '', corrected_value.lower())

            cell = worksheet.cell(row=row, column=col_idx[column])
            if original_clean == corrected_clean:
                cell.fill = green_fill  # Запись осталась неизменной (ошибок не обнаружено)
            elif corrected_value != original_value and any(word.lower() not in correct_words for word in re.findall(r'\w+', corrected_clean)):
                cell.fill = yellow_fill  # Встречено неизвестное слово при исправлении записи
            else:
                cell.fill = red_fill  # Редактирование записи прошло успешно
                cell.value = original_value  # Восстанавливаем исходное значение для красных ячеек

    workbook.save(output_file1 if 'Профессия' in columns else output_file2)

# Раскрашиваем ячейки и восстанавливаем исходные значения для "Потребность персонала"
wb1 = load_workbook(output_file1)
ws1 = wb1.active
color_and_restore_cells(wb1, ws1, df1_corrected, ['Профессия'])

# Раскрашиваем ячейки и восстанавливаем исходные значения для "ЦЗН"
wb2 = load_workbook(output_file2)
ws2 = wb2.active
color_and_restore_cells(wb2, ws2, df2_corrected, ['Должность', 'Специальность'])

print(f"Исправленные файлы сохранены с раскрашенными ячейками как {output_file1} и {output_file2}")